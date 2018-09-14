# -*- coding: utf-8 -*-
"""
Created on Fri Dec  8 02:30:14 2017

@author: Shomi Nanwani
"""

import requests
from time import localtime, strftime
import sys
import os
import check_internet
import openpyxl
import config
import time


def cook(cj):
    j=str(cj)
    t2=j.find(' for ')
    t1=int(j.find('~'))+1
    tokken=str(j[t1:t2])
    return tokken 

def check_message_len(message):
    global msglen
    message=message.replace(' ', '+')
    msglen=140-len(message)
    if len(message)>140:
        print('Message length is greater than 140')
        sys.exit(0)

def current_time(expression):
    if expression=='date':
        tme=strftime("%d-%m", localtime())
    elif expression=='time':
        tme=strftime("%H:%M:%S", localtime())
    elif expression=='date-time':
        tme=strftime("%d-%m//%H:%M:%S", localtime())
    elif expression=='time-date':
        tme=strftime("%H:%M:%S//%d-%m", localtime())
    return tme
    
def send_sms(mobile_number,password,text):
    global token,message
    s=requests.Session()
    header_payload={'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8', 'Connection': 'close'}
    login_payload={'mobileNo':mobile_number,'password':password,'CatType':''}
    s.get('http://www.way2sms.com', allow_redirects=True, timeout=10)
    resp = s.post('http://www.way2sms.com/re-login', headers=header_payload, data=login_payload, allow_redirects=True, timeout=10)
    if resp.text=='send-sms':
        print("Login Successful.")
    else:
        print("Login Failed.")
        sys.exit(0)
    token=cook(s.cookies)
    message=text+''+current_time('time-date')+'\nID: '+token
    sms_payload={'Token':token, 'message':message, 'toMobile':config.receiver_number_windows_state_alert[0], 'ssaction':'ss'}
    response=s.post('http://www.way2sms.com/smstoss', data=sms_payload, allow_redirects=True, timeout=10)
    data=response.text.encode("utf-8")
    if "6" in str(data):
        msg='Daily quota finished.'
    elif "0" in str(data):
        msg='Message Sent.'
    elif "5" in str(data):
        msg='SPAM Words'
    else:
        msg='Error Occured'
        error_log(msg,str(data),mobile_number)
    return msg

def error_log(msg,data,mobile_number):
	f = open('windows_state_change_alert_errors.txt', 'a')
	f.write(current_time('date-time')+':'+mobile_number+':'+msg+'\n\n'+str(data))
	f.write('**************************************************************************************************************************')
	f.close

def update_excel_data(status,msg_sent,token,registered_mobile,msg,err_msg,err_logged):
    path=config.path_windows_state_alert
    wb = openpyxl.load_workbook(path)
    sheet_names=wb.sheetnames
    ws = wb[sheet_names[0]]
    ws.cell(row=1, column=1).value = 'Date'
    ws.cell(row=1, column=2).value = 'Time'
    ws.cell(row=1, column=3).value = 'Windows Status'
    ws.cell(row=1, column=4).value = 'Alert Msg Sent'
    ws.cell(row=1, column=5).value = 'Alert Msg Time'
    ws.cell(row=1, column=6).value = 'Token'
    ws.cell(row=1, column=7).value = 'Sent Via'
    ws.cell(row=1, column=8).value = 'Message'
    ws.cell(row=1, column=9).value = 'Error'
    ws.cell(row=1, column=10).value = 'Error Output to a File'
    wb.save(path)
	
    try:
        windows_status = ([str(ws[x][2].value) for x in range(1,ws.max_row+1)])
    except IndexError:
        windows_status = []
		
    ws.cell(row=len(windows_status)+1, column=1).value = current_time('date')
    ws.cell(row=len(windows_status)+1, column=2).value = current_time('time')
    ws.cell(row=len(windows_status)+1, column=3).value = status
    ws.cell(row=len(windows_status)+1, column=4).value = msg_sent
    ws.cell(row=len(windows_status)+1, column=5).value = current_time('date-time')
    ws.cell(row=len(windows_status)+1, column=6).value = token
    ws.cell(row=len(windows_status)+1, column=7).value = registered_mobile
    ws.cell(row=len(windows_status)+1, column=8).value = msg
    if err_msg=='':
        ws.cell(row=len(windows_status)+1, column=9).value = ''
    else:
        ws.cell(row=len(windows_status)+1, column=9).value = registered_mobile+' '+err_msg
    ws.cell(row=len(windows_status)+1, column=10).value = err_logged
    wb.save(path)


if len (sys.argv) != 2 :
    print("Usage: python windows_state_change_alert.py Logged_In|hibernated|started|sleep")
    sys.exit (1)

status=sys.argv[1]
message='Laptop State: '+status+'\nSate changed at: '+current_time('time-date')+'\nMessage sent at: '
check_message_len(message)

x=0
msg='x'
while x<len(config.registered_mobile_windows_state_alert):
    send_msg=message
    if check_internet.is_connected() is True:
        msg=send_sms(config.registered_mobile_windows_state_alert[x],config.pass_windows_state_alert[x],send_msg)
        print(msg)
    else:
        update_excel_data(status,'0','',config.registered_mobile_windows_state_alert[x],'','Internet Disconnected','0')
        print('Internet Disconnected')
        time.sleep(600)
        # sys.exit(0)
    if msg=='Message Sent.':
        update_excel_data(status,'1',token,config.registered_mobile_windows_state_alert[x],send_msg,'','0')
        sys.exit(0)
    elif msg=='Error Occured':
        err_msg='Error Occured'
        update_excel_data(status,'0',token,config.registered_mobile_windows_state_alert[x],send_msg,err_msg,'1')
        x+=1
    elif msg=='Daily quota finished.' or msg=='SPAM Words' or msg=='Receiver Blocked this Number.' or msg=='Login Failed' or msg=='Unauthorized Login':
        err_msg='Error: '+msg
        update_excel_data(status,'0',token,config.registered_mobile_windows_state_alert[x],send_msg,err_msg,'0')
        x+=1
